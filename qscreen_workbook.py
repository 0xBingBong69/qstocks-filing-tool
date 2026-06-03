#!/usr/bin/env python3
"""qscreen_workbook.py — a model-ready Excel transcript of a QSE filing.

Turns one extracted filing (and optionally several, for more years) into a
multi-sheet .xlsx an analyst can drop straight into a model:
  • Summary               — company / period / currency / unit / auditor
  • one sheet per statement (Income Statement, Balance Sheet, …) as printed,
    with the current + comparative period columns and numeric value cells
  • Financials (multi-year) — canonical metrics as rows, fiscal years as columns
  • Segments / Notes        — when the filing has them

    from qscreen_workbook import save_workbook, workbook_bytes
    save_workbook(filing, "QNBK_2023_FY_filing.xlsx")          # one filing
    save_workbook(filing, "QNBK.xlsx", filings=[f2022, f2023]) # more years

Numbers are written as numeric cells (so Excel can sum/link); the qscreen.app
JSON contract is untouched — this is a parallel, human-facing output.
"""
from __future__ import annotations

import argparse
import io
import json
import re
import sys
from pathlib import Path

from qscreen_series import build_series

_STATEMENT_SHEET = {
    "income_statement": "Income Statement",
    "balance_sheet": "Balance Sheet",
    "cash_flow": "Cash Flow",
    "changes_in_equity": "Changes in Equity",
    "comprehensive_income": "Comprehensive Income",
}
# Group the multi-year grid in reading order: P&L, balance sheet, cash flow, …
_PREFIX_ORDER = ["IS_", "BS_", "CF_", "SCE_", "EQ_", "OCI_", "KPI_"]


def _num(x):
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return x
    if isinstance(x, str):
        t = x.strip().replace(",", "").replace("(", "-").replace(")", "")
        try:
            return float(t)
        except ValueError:
            return None
    return None


def _unit_note(scale):
    return {1: "actual (×1)", 1000: "thousands (×1,000)",
            1000000: "millions (×1,000,000)"}.get(scale, f"×{scale}" if scale else "actual")


def _import_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import Font
        return openpyxl, Font
    except Exception:
        raise SystemExit("xlsx workbook needs openpyxl:  pip install openpyxl")


def _safe_sheet_name(name, used: set) -> str:
    name = re.sub(r"[:\\/?*\[\]]", " ", str(name)).strip()[:31] or "Sheet"
    base, i = name, 2
    while name in used:
        suffix = f" ({i})"
        name = (base[:31 - len(suffix)] + suffix)
        i += 1
    used.add(name)
    return name


def _prior_period_labels(st: dict) -> list:
    priors: list = []
    for li in st.get("line_items") or []:
        for c in (li.get("comparatives") or []):
            pl = c.get("period_label") if isinstance(c, dict) else None
            if pl and pl not in priors:
                priors.append(pl)
    return priors


def _ordered_codes(codes: list) -> list:
    def key(c):
        for i, p in enumerate(_PREFIX_ORDER):
            if c.startswith(p):
                return (i, c)
        return (len(_PREFIX_ORDER), c)
    return sorted(codes, key=key)


def _bold_row(ws, row_idx, ncols, Font):
    for c in range(1, ncols + 1):
        ws.cell(row_idx, c).font = Font(bold=True)


def _summary_sheet(ws, filing, Font):
    meta, audit = filing.get("metadata") or {}, filing.get("audit") or {}
    ws.append(["QScreen — financial transcript"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for k, v in [
        ("Company", meta.get("company_name")), ("Symbol", meta.get("symbol")),
        ("Sector", meta.get("sector")), ("Fiscal year", meta.get("fiscal_year")),
        ("Period", meta.get("fiscal_period")), ("Period end", meta.get("period_end")),
        ("Currency", meta.get("currency")), ("Figures in", _unit_note(meta.get("unit_scale"))),
        ("Reporting framework", meta.get("reporting_framework")),
        ("Consolidated", meta.get("consolidated")),
        ("Auditor", audit.get("auditor_name")), ("Audit opinion", audit.get("opinion_type")),
        ("Source file", meta.get("source_file")), ("Extracted at", meta.get("extracted_at")),
    ]:
        ws.append([k, v])
        ws.cell(ws.max_row, 1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48


def _statement_sheet(ws, st, Font):
    priors = _prior_period_labels(st)
    header = ["Label", "Code", "Note", st.get("period_label") or "Current"] + priors
    ws.append(header)
    _bold_row(ws, 1, len(header), Font)
    for li in st.get("line_items") or []:
        depth = int(li.get("depth") or 0)
        comp = {c.get("period_label"): _num(c.get("value"))
                for c in (li.get("comparatives") or []) if isinstance(c, dict)}
        ws.append(["   " * depth + str(li.get("label_verbatim") or ""),
                   li.get("account_code"), li.get("note_ref"), _num(li.get("value"))]
                  + [comp.get(p) for p in priors])
        if li.get("is_subtotal"):
            _bold_row(ws, ws.max_row, len(header), Font)
    ws.column_dimensions["A"].width = 46


def _multiyear_sheet(ws, series, Font):
    years = sorted(series.get("years") or {})
    ws.append(["Metric", "Code"] + years)
    _bold_row(ws, 1, 2 + len(years), Font)

    def label_for(code):
        for y in reversed(years):
            lbl = (series["years"][y].get("labels") or {}).get(code)
            if lbl:
                return lbl
        return code
    for code in _ordered_codes(series.get("codes") or []):
        ws.append([label_for(code), code]
                  + [_num((series["years"][y].get("metrics") or {}).get(code)) for y in years])
    ws.column_dimensions["A"].width = 42


def _segments_sheet(ws, segs, Font):
    mkeys = sorted({k for sg in segs for k in (sg.get("metrics") or {})})
    header = ["Dimension", "Segment", "Currency", "Period"] + mkeys
    ws.append(header)
    _bold_row(ws, 1, len(header), Font)
    for sg in segs:
        m = sg.get("metrics") or {}
        ws.append([sg.get("dimension"), sg.get("name"), sg.get("currency"), sg.get("period_label")]
                  + [_num(m.get(k)) for k in mkeys])
    ws.column_dimensions["B"].width = 26


def _notes_sheet(ws, notes, Font):
    ws.append(["Number", "Title", "Category", "Text"])
    _bold_row(ws, 1, 4, Font)
    for nt in notes:
        ws.append([nt.get("number"), nt.get("title"), nt.get("category"), nt.get("verbatim_text")])
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["D"].width = 90


def build_workbook(filing: dict, filings: list[dict] | None = None):
    """Return an openpyxl Workbook with Summary + per-statement + multi-year
    (+ Segments/Notes) sheets."""
    openpyxl, Font = _import_openpyxl()
    wb = openpyxl.Workbook()
    used: set = set()

    ws = wb.active
    ws.title = _safe_sheet_name("Summary", used)
    _summary_sheet(ws, filing, Font)

    for st in filing.get("statements") or []:
        title = _STATEMENT_SHEET.get(st.get("type")) or (st.get("title") or st.get("type") or "Statement")
        _statement_sheet(wb.create_sheet(_safe_sheet_name(title, used)), st, Font)

    sym = (filing.get("metadata") or {}).get("symbol") or "SERIES"
    series = build_series(sym, filings or [filing])
    if series.get("years"):
        _multiyear_sheet(wb.create_sheet(_safe_sheet_name("Financials (multi-year)", used)), series, Font)

    if filing.get("segments"):
        _segments_sheet(wb.create_sheet(_safe_sheet_name("Segments", used)), filing["segments"], Font)
    if filing.get("notes"):
        _notes_sheet(wb.create_sheet(_safe_sheet_name("Notes", used)), filing["notes"], Font)
    return wb


def save_workbook(filing: dict, path: str, filings: list[dict] | None = None) -> str:
    build_workbook(filing, filings).save(path)
    return path


def workbook_bytes(filing: dict, filings: list[dict] | None = None) -> bytes:
    buf = io.BytesIO()
    build_workbook(filing, filings).save(buf)
    return buf.getvalue()


def main() -> int:
    p = argparse.ArgumentParser(
        description="Build an Excel financial-transcript workbook from one or more filings "
                    "(same company; extra years extend the multi-year grid).")
    p.add_argument("filings", nargs="+", help="SYMBOL_YEAR_PERIOD_filing.json files")
    p.add_argument("--symbol", help="ticker for the output filename (else taken from the latest filing)")
    p.add_argument("--out", help="output .xlsx path (default: <SYMBOL>_transcript.xlsx)")
    args = p.parse_args()

    filings = [json.loads(Path(fp).read_text(encoding="utf-8")) for fp in args.filings]
    filings.sort(key=lambda f: (f.get("metadata") or {}).get("fiscal_year") or 0)
    primary = filings[-1]                                   # latest year supplies the statement sheets
    sym = (args.symbol or (primary.get("metadata") or {}).get("symbol") or "filing").upper()
    out = args.out or f"{sym}_transcript.xlsx"
    save_workbook(primary, out, filings)
    years = sorted({(f.get("metadata") or {}).get("fiscal_year") for f in filings
                    if (f.get("metadata") or {}).get("fiscal_year")})
    print(f"📑 Excel transcript ({len(filings)} filing(s), reported years {years}) → {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())

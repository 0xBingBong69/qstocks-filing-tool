"""Offline test suite for the QScreen filing tool — no network, no API key.

Covers the contract validator, the LLM-output normalizer, the lossless
cross-window merge, page windowing, the hardened JSON parser, exports, the
batch manifest reader, OCR fall-through, and the LLM/upload HTTP wrappers
(with requests stubbed).
"""
from __future__ import annotations

import json
from types import SimpleNamespace

import pytest

import qscreen_ingest as e


# ── fixtures / helpers ───────────────────────────────────────────────────────

def good_filing() -> dict:
    f = e.empty_filing()
    f["metadata"].update({"symbol": "QNBK", "sector": "conventional_bank",
                          "fiscal_year": 2023, "fiscal_period": "FY", "unit_scale": 1000})
    f["audit"].update({"opinion_type": "unqualified", "verbatim_text": "In our opinion …"})
    f["statements"].append({"type": "income_statement", "title": "Income", "period_label": "2023",
                            "verbatim_text": "NII 1",
                            "line_items": [{"account_code": "IS_NET_INTEREST", "label_verbatim": "NII",
                                            "value": 1, "note_ref": "24", "depth": 0, "is_subtotal": False}]})
    f["notes"].append({"number": "27", "title": "Contingencies", "category": "contingent_liabilities",
                       "structured": {}, "verbatim_text": "…"})
    return f


def llm_args(**over):
    base = dict(symbol="QNBK", sector="conventional_bank", year=2023, period="FY",
                provider="openrouter", base_url=None, model=None, max_tokens=128,
                timeout=5, retries=3, no_json_mode=False, llm_key="sk-test",
                pages_per_chunk=12, overlap=1, no_chunk=False)
    base.update(over)
    return SimpleNamespace(**base)


# ── contract validation ──────────────────────────────────────────────────────

def test_validate_accepts_good():
    assert e.validate_filing(good_filing()) == []


def test_validate_rejects_empty_statements():
    f = good_filing()
    f["statements"] = []
    assert any("statements: empty" in p for p in e.validate_filing(f))


def test_validate_rejects_unknown_opinion():
    f = good_filing()
    f["audit"]["opinion_type"] = "great"
    assert any("opinion_type" in p for p in e.validate_filing(f))


def test_validate_flags_opinion_without_verbatim():
    f = good_filing()
    f["audit"]["verbatim_text"] = ""
    assert any("verbatim_text" in p for p in e.validate_filing(f))


def test_validate_flags_unknown_account_code():
    f = good_filing()
    f["statements"][0]["line_items"][0]["account_code"] = "IS_MADE_UP"
    assert any("unknown" in p for p in e.validate_filing(f))


def test_validate_flags_bad_unit_scale():
    f = good_filing()
    f["metadata"]["unit_scale"] = 100
    assert any("unit_scale" in p for p in e.validate_filing(f))


def test_validate_flags_lossy_line_item():
    f = good_filing()
    f["statements"][0]["line_items"][0]["label_verbatim"] = ""
    assert any("label_verbatim" in p for p in e.validate_filing(f))


# ── normalization ────────────────────────────────────────────────────────────

def test_normalize_aliases_and_unknown_code():
    drifted = {
        "metadata": {"ticker": "QIBK", "company": "Qatar Islamic Bank", "sector": "Islamic Bank",
                     "reporting_currency": "QAR", "framework": "AAOIFI", "unit_scale": 1000},
        "audit": {"opinion_type": "unqualified", "opinion_text": "In our opinion …",
                  "key_audit_matters": [{"title": "ECL", "description": "judgemental ECL"}],
                  "emphasis_of_matter": "one matter"},
        "statements": [{"type": "income_statement", "period": "year_ended_2024", "verbatim_text": "…",
                        "line_items": [{"label_verbatim": "x", "value": 1},
                                       {"account_code": "IS_OTHER_COMPREHENSIVE_INCOME",
                                        "label_verbatim": "OCI", "value": 2}]}],
        "notes": [], "extraction_quality": {},
    }
    n = e.normalize_filing(drifted)
    assert n["metadata"]["symbol"] == "QIBK"
    assert n["metadata"]["company_name"] == "Qatar Islamic Bank"
    assert n["metadata"]["sector"] == "islamic_bank"
    assert n["metadata"]["currency"] == "QAR"
    assert n["metadata"]["reporting_framework"] == "AAOIFI"
    assert n["audit"]["verbatim_text"]                       # opinion_text → verbatim_text
    assert n["audit"]["key_audit_matters"][0]["text"] == "judgemental ECL"
    assert n["audit"]["emphasis_of_matter"] == ["one matter"]  # str coerced to list
    assert n["statements"][0]["period_label"] == "year_ended_2024"
    assert n["statements"][0]["line_items"][1]["account_code"] is None  # unknown → null
    assert any("IS_OTHER_COMPREHENSIVE_INCOME" in u
               for u in n["extraction_quality"]["unmapped_labels"])


@pytest.mark.parametrize("raw,expect", [
    ("Islamic Bank", "islamic_bank"), ("commercial-bank", "conventional_bank"),
    ("Takaful", "insurance"), ("Petrochemicals industrials", "industrial"),
    ("Real Estate", "other"), ("", None),
])
def test_normalize_sector(raw, expect):
    assert e._normalize_sector(raw) == expect


@pytest.mark.parametrize("raw,expect", [
    (1000, 1000), ("in thousands", 1000), ("QAR millions", 1000000), ("nonsense", None),
])
def test_normalize_unit_scale(raw, expect):
    assert e._normalize_unit_scale(raw) == expect


# ── lossless cross-window merge ──────────────────────────────────────────────

def test_merge_combines_windows():
    a = e.empty_filing(); a["audit"].update({"opinion_type": "unqualified", "verbatim_text": "op …"})
    b = e.empty_filing(); b["statements"].append(
        {"type": "balance_sheet", "verbatim_text": "BS …",
         "line_items": [{"label_verbatim": "Total assets", "value": 9, "account_code": "BS_TOTAL_ASSETS"}]})
    c = e.empty_filing(); c["notes"].append(
        {"number": "5", "title": "Sukuk", "category": "sukuk_islamic", "structured": {}, "verbatim_text": "sukuk …"})
    m = e.merge_filings([a, b, c])
    assert m["audit"]["opinion_type"] == "unqualified"
    assert m["statements"] and m["notes"]


def test_merge_unions_split_statement_lossless():
    w1 = e.empty_filing(); w2 = e.empty_filing()
    w1["statements"].append({"type": "income_statement", "title": "IS", "period_label": "2024",
        "verbatim_text": "the much longer first-window verbatim block",
        "line_items": [
            {"label_verbatim": "Revenue", "value": 10, "account_code": "IS_REVENUE"},
            {"label_verbatim": "Net interest", "value": 5, "account_code": None}]})
    w2["statements"].append({"type": "income_statement", "title": "IS", "period_label": "2024",
        "verbatim_text": "short",
        "line_items": [
            {"label_verbatim": "Net interest", "value": 5, "account_code": "IS_NET_INTEREST"},  # overlap dup
            {"label_verbatim": "Net income", "value": 3, "account_code": "IS_NET_INCOME"}]})     # split row
    st = e.merge_filings([w1, w2])["statements"][0]
    labels = [li["label_verbatim"] for li in st["line_items"]]
    assert labels == ["Revenue", "Net interest", "Net income"]          # no loss, no dup
    ni = next(li for li in st["line_items"] if li["label_verbatim"] == "Net interest")
    assert ni["account_code"] == "IS_NET_INTEREST"                      # null upgraded from dup
    assert st["verbatim_text"].startswith("the much longer")           # longest verbatim kept


def test_merge_dedups_kams_and_aggregates_quality():
    w1 = e.empty_filing()
    w1["audit"].update({"opinion_type": "unqualified", "verbatim_text": "opinion",
                        "key_audit_matters": [{"title": "ECL", "text": "aaa"}]})
    w1["extraction_quality"] = {"confidence": 0.9, "warnings": ["w1"], "unmapped_labels": ["u1"]}
    w2 = e.empty_filing()
    w2["audit"].update({"key_audit_matters": [{"title": "ECL", "text": "aaa"}]})  # duplicate KAM
    w2["extraction_quality"] = {"confidence": 0.5, "warnings": ["w1", "w2"], "unmapped_labels": ["u2"]}
    m = e.merge_filings([w1, w2])
    assert len(m["audit"]["key_audit_matters"]) == 1
    assert m["extraction_quality"]["confidence"] == 0.5                 # min
    assert m["extraction_quality"]["warnings"] == ["w1", "w2"]          # dedup + sorted
    assert m["extraction_quality"]["unmapped_labels"] == ["u1", "u2"]


# ── page windowing ───────────────────────────────────────────────────────────

def test_page_windows_overlap_and_coverage():
    pages = [{"num": i, "text": str(i)} for i in range(1, 8)]   # 7 pages
    wins = e.page_windows(pages, size=3, overlap=1)             # step 2, stops at the last full window
    assert [(w[0]["num"], w[-1]["num"]) for w in wins] == [(1, 3), (3, 5), (5, 7)]


def test_page_windows_no_chunk_when_size_zero():
    pages = [{"num": i, "text": str(i)} for i in range(1, 5)]
    assert e.page_windows(pages, size=0, overlap=0) == [pages]


# ── hardened JSON parsing ────────────────────────────────────────────────────

def test_parse_plain_object():
    assert e.parse_llm_json('{"a": 1}') == {"a": 1}


def test_parse_strips_code_fences():
    assert e.parse_llm_json('```json\n{"a": 1}\n```') == {"a": 1}


def test_parse_tolerates_trailing_prose_and_inner_braces():
    assert e.parse_llm_json('{"a": "x}y", "b": {"n": 2}} thanks!')["b"]["n"] == 2


def test_parse_handles_escaped_quotes():
    assert e.parse_llm_json(r'{"a": "she said \"hi\" {"}')["a"] == 'she said "hi" {'


def test_parse_no_object_raises():
    with pytest.raises(ValueError):
        e.parse_llm_json("no json here")


# ── exports ──────────────────────────────────────────────────────────────────

def test_flatten_and_csv_export(tmp_path):
    f = good_filing()
    rows = e.flatten_line_items(f)
    assert rows and rows[0]["account_code"] == "IS_NET_INTEREST"
    out = tmp_path / "x.csv"
    n = e.export_csv(f, str(out))
    lines = out.read_text(encoding="utf-8").splitlines()
    assert n == 1
    assert lines[0].split(",")[0] == "statement_type"


def test_xlsx_export(tmp_path):
    out = tmp_path / "x.xlsx"
    n = e.export_xlsx(good_filing(), str(out))
    from openpyxl import load_workbook
    ws = load_workbook(str(out)).active
    assert n == 1
    assert ws.max_row == 2 and ws.max_column == len(e.EXPORT_COLUMNS)


# ── batch manifest ───────────────────────────────────────────────────────────

def test_read_manifest_ok(tmp_path):
    p = tmp_path / "m.csv"
    p.write_text("pdf,symbol,sector,year,period\na.pdf,QIBK,Islamic Bank,2024,FY\n"
                 "b.pdf,QNBK,conventional_bank,2023,\n", encoding="utf-8")
    rows = e.read_manifest(str(p))
    assert [r["symbol"] for r in rows] == ["QIBK", "QNBK"]


def test_read_manifest_missing_column(tmp_path):
    p = tmp_path / "m.csv"
    p.write_text("pdf,symbol,year\na.pdf,QIBK,2024\n", encoding="utf-8")
    with pytest.raises(SystemExit):
        e.read_manifest(str(p))


def test_read_manifest_empty(tmp_path):
    p = tmp_path / "m.csv"
    p.write_text("pdf,symbol,sector,year\n", encoding="utf-8")
    with pytest.raises(SystemExit):
        e.read_manifest(str(p))


# ── OCR fall-through (pdfplumber stubbed; OCR deps absent) ────────────────────

class _FakePage:
    def __init__(self, text): self._t = text
    def extract_text(self): return self._t
    def extract_tables(self): return []


class _FakePDF:
    def __init__(self, texts): self.pages = [_FakePage(t) for t in texts]
    def __enter__(self): return self
    def __exit__(self, *a): return False


def _stub_pdf(monkeypatch, texts, tmp_path):
    import pdfplumber
    monkeypatch.setattr(pdfplumber, "open", lambda path: _FakePDF(texts))
    pdf = tmp_path / "f.pdf"
    pdf.write_bytes(b"%PDF-1.4 stub")
    return str(pdf)


def test_ocr_auto_warns_when_unavailable(monkeypatch, tmp_path, capsys):
    path = _stub_pdf(monkeypatch, ["full page of text " * 5, ""], tmp_path)
    monkeypatch.setattr(e, "_ocr_pages", lambda p, nums: (_ for _ in ()).throw(e.OcrUnavailable("no deps")))
    pages, sha = e.pdf_to_pages(path, ocr_mode="auto")
    assert len(pages) == 2 and pages[1]["text"] == ""        # unchanged, no crash
    assert "likely" in capsys.readouterr().out.lower()
    assert len(sha) == 64


def test_ocr_always_recovers_text(monkeypatch, tmp_path):
    path = _stub_pdf(monkeypatch, ["", ""], tmp_path)
    monkeypatch.setattr(e, "_ocr_pages", lambda p, nums: {n: f"ocr-page-{n}" for n in nums})
    pages, _ = e.pdf_to_pages(path, ocr_mode="always")
    assert "ocr-page-1" in pages[0]["text"] and "ocr-page-2" in pages[1]["text"]


def test_ocr_always_raises_when_unavailable(monkeypatch, tmp_path):
    path = _stub_pdf(monkeypatch, [""], tmp_path)
    monkeypatch.setattr(e, "_ocr_pages", lambda p, nums: (_ for _ in ()).throw(e.OcrUnavailable("no deps")))
    with pytest.raises(SystemExit):
        e.pdf_to_pages(path, ocr_mode="always")


# ── extract_filing orchestration (call_llm stubbed) ──────────────────────────

def test_extract_single_pass(monkeypatch):
    payload = json.dumps({
        "metadata": {"ticker": "QNBK"},
        "audit": {"opinion_type": "unqualified", "verbatim_text": "op"},
        "statements": [{"type": "balance_sheet", "verbatim_text": "BS",
                        "line_items": [{"label_verbatim": "Total assets", "value": 1,
                                        "account_code": "BS_TOTAL_ASSETS"}]}],
        "notes": [], "extraction_quality": {"confidence": 0.8},
    })
    monkeypatch.setattr(e, "call_llm", lambda messages, args: payload)
    pages = [{"num": 1, "text": "page one"}]
    out = e.extract_filing(pages, llm_args(no_chunk=True))
    assert out["metadata"]["symbol"] == "QNBK"               # normalized alias
    assert out["statements"][0]["type"] == "balance_sheet"


def test_extract_windowed_merges(monkeypatch):
    def fake(messages, args):
        # Different statement per window so the merge must combine them.
        body = messages[1]["content"]
        if "PAGE 1" in body:
            return json.dumps({"metadata": {}, "audit": {"opinion_type": "unknown", "verbatim_text": ""},
                               "statements": [{"type": "income_statement", "verbatim_text": "IS",
                                               "line_items": [{"label_verbatim": "Rev", "value": 1,
                                                               "account_code": "IS_REVENUE"}]}],
                               "notes": [], "extraction_quality": {}})
        return json.dumps({"metadata": {}, "audit": {"opinion_type": "unknown", "verbatim_text": ""},
                           "statements": [{"type": "balance_sheet", "verbatim_text": "BS",
                                           "line_items": [{"label_verbatim": "TA", "value": 2,
                                                           "account_code": "BS_TOTAL_ASSETS"}]}],
                           "notes": [], "extraction_quality": {}})
    monkeypatch.setattr(e, "call_llm", fake)
    pages = [{"num": i, "text": f"content {i}"} for i in range(1, 6)]
    out = e.extract_filing(pages, llm_args(no_chunk=False, pages_per_chunk=2, overlap=1))
    types = {s["type"] for s in out["statements"]}
    assert {"income_statement", "balance_sheet"} <= types


def test_extract_windowed_skips_unparseable(monkeypatch):
    calls = {"n": 0}
    def fake(messages, args):
        calls["n"] += 1
        if calls["n"] == 1:
            return "not json at all"
        return json.dumps({"metadata": {}, "audit": {"opinion_type": "unknown", "verbatim_text": ""},
                           "statements": [{"type": "cash_flow", "verbatim_text": "CF",
                                           "line_items": [{"label_verbatim": "OCF", "value": 1,
                                                           "account_code": "CF_OCF"}]}],
                           "notes": [], "extraction_quality": {}})
    monkeypatch.setattr(e, "call_llm", fake)
    pages = [{"num": i, "text": f"c{i}"} for i in range(1, 6)]
    out = e.extract_filing(pages, llm_args(no_chunk=False, pages_per_chunk=2, overlap=1))
    assert any(s["type"] == "cash_flow" for s in out["statements"])  # survived a bad window


# ── provider registry / resolution (no network) ─────────────────────────────

_PROVIDER_ENV = ("MINIMAX_API_KEY", "OPENROUTER_API_KEY", "OPENAI_API_KEY",
                 "ANTHROPIC_API_KEY", "MOONSHOT_API_KEY", "KIMI_API_KEY",
                 "QSCREEN_PROVIDER", "LLM_PROVIDER", "QSCREEN_MODEL", "LLM_API_KEY")


@pytest.fixture
def clean_provider_env(monkeypatch):
    for k in _PROVIDER_ENV:
        monkeypatch.delenv(k, raising=False)


def _pargs(**over):
    base = dict(provider=None, base_url=None, model=None, llm_key=None,
                max_tokens=128, no_json_mode=False, retries=2, timeout=5)
    base.update(over)
    return SimpleNamespace(**base)


def test_canonical_provider_aliases():
    assert e.canonical_provider("claude") == "anthropic"
    assert e.canonical_provider("CLAUDE") == "anthropic"
    assert e.canonical_provider("moonshot") == "kimi"
    assert e.canonical_provider("gpt") == "openai"
    assert e.canonical_provider(None) is None


def test_default_model_and_listing():
    assert e.default_model("minimax") == "MiniMax-M2"
    assert e.default_model("claude") == e.PROVIDERS["anthropic"]["default_model"]
    assert e.default_model("nope") is None
    txt = e.list_providers()
    for name in ("minimax", "openrouter", "kimi", "openai", "anthropic", "custom"):
        assert name in txt


def test_provider_base_urls_and_key_links():
    # Verified against each provider's docs (June 2026). Kimi must be the GLOBAL
    # endpoint — the .cn endpoint silently 401s international keys.
    assert e.PROVIDERS["minimax"]["base_url"] == "https://api.minimax.io/v1"
    assert e.PROVIDERS["kimi"]["base_url"] == "https://api.moonshot.ai/v1"
    assert e.PROVIDERS["openrouter"]["base_url"] == "https://openrouter.ai/api/v1"
    assert e.PROVIDERS["openai"]["base_url"] == "https://api.openai.com/v1"
    assert e.PROVIDERS["anthropic"]["base_url"] == "https://api.anthropic.com/v1"
    for name, p in e.PROVIDERS.items():
        assert p["key_url"].startswith("https://"), name
        assert p["label"], name
    txt = e.list_providers()
    assert "https://platform.moonshot.ai/console/api-keys" in txt
    assert "https://console.anthropic.com/settings/keys" in txt


def test_resolve_explicit_minimax(clean_provider_env):
    cfg = e.resolve_provider(_pargs(provider="minimax", llm_key="sk-mm"))
    assert cfg["name"] == "minimax" and cfg["kind"] == "openai"
    assert cfg["base_url"] == "https://api.minimax.io/v1" and cfg["key"] == "sk-mm"


def test_resolve_claude_alias_is_anthropic(clean_provider_env):
    cfg = e.resolve_provider(_pargs(provider="claude", llm_key="sk-an"))
    assert cfg["name"] == "anthropic" and cfg["kind"] == "anthropic"


def test_resolve_autodetects_from_env_key(clean_provider_env, monkeypatch):
    monkeypatch.setenv("OPENAI_API_KEY", "sk-oai")
    cfg = e.resolve_provider(_pargs())
    assert cfg["name"] == "openai" and cfg["key"] == "sk-oai"


def test_resolve_env_provider_and_model_override(clean_provider_env, monkeypatch):
    monkeypatch.setenv("QSCREEN_PROVIDER", "anthropic")
    monkeypatch.setenv("ANTHROPIC_API_KEY", "sk-an")
    monkeypatch.setenv("QSCREEN_MODEL", "claude-custom")
    cfg = e.resolve_provider(_pargs())
    assert cfg["name"] == "anthropic" and cfg["model"] == "claude-custom"


def test_resolve_custom_requires_base_url(clean_provider_env):
    with pytest.raises(SystemExit):
        e.resolve_provider(_pargs(provider="custom", llm_key="k"))
    cfg = e.resolve_provider(_pargs(provider="custom", base_url="https://x/v1", model="m", llm_key="k"))
    assert cfg["base_url"] == "https://x/v1" and cfg["kind"] == "openai"


def test_resolve_missing_key_raises(clean_provider_env):
    with pytest.raises(SystemExit) as ei:
        e.resolve_provider(_pargs(provider="openai"))
    assert "OPENAI_API_KEY" in str(ei.value)


def test_anthropic_request_shape():
    msgs = [{"role": "system", "content": "SYS"}, {"role": "user", "content": "U"}]
    cfg = {"name": "anthropic", "base_url": "https://api.anthropic.com/v1",
           "kind": "anthropic", "model": "claude-x", "key": "k"}
    url, headers, payload, extract = e._anthropic_request(msgs, cfg, _pargs(max_tokens=99))
    assert url.endswith("/messages")
    assert headers["x-api-key"] == "k" and headers["anthropic-version"] == "2023-06-01"
    assert payload["system"] == "SYS" and payload["max_tokens"] == 99
    assert payload["messages"][-1] == {"role": "assistant", "content": "{"}   # JSON prefill
    assert extract({"content": [{"type": "text", "text": '"a": 1}'}]}) == '{"a": 1}'


def test_openai_request_shape():
    msgs = [{"role": "system", "content": "S"}, {"role": "user", "content": "U"}]
    cfg = {"name": "openai", "base_url": "https://api.openai.com/v1",
           "kind": "openai", "model": "gpt", "key": "k"}
    url, headers, payload, extract = e._openai_request(msgs, cfg, _pargs())
    assert url.endswith("/chat/completions") and headers["Authorization"] == "Bearer k"
    assert payload["response_format"] == {"type": "json_object"}
    assert extract({"choices": [{"message": {"content": "X"}}]}) == "X"


def test_call_llm_anthropic_end_to_end(clean_provider_env, monkeypatch):
    import requests
    monkeypatch.setattr(requests, "post",
                        lambda *a, **k: _Resp(200, {"content": [{"type": "text", "text": '"ok": 1}'}]}))
    out = e.call_llm([{"role": "system", "content": "S"}, {"role": "user", "content": "U"}],
                     _pargs(provider="anthropic", llm_key="k"))
    assert out == '{"ok": 1}'                                   # opening brace reconstructed


# ── HTTP wrappers (requests stubbed) ─────────────────────────────────────────

class _Resp:
    def __init__(self, status=200, payload=None, text=""):
        self.status_code = status
        self._payload = payload or {}
        self.text = text
    def json(self): return self._payload
    def raise_for_status(self):
        if self.status_code >= 400:
            raise RuntimeError(f"HTTP {self.status_code}")


def test_call_llm_success(monkeypatch):
    import requests
    ok = _Resp(200, {"choices": [{"message": {"content": '{"ok": 1}'}}]})
    monkeypatch.setattr(requests, "post", lambda *a, **k: ok)
    assert e.call_llm([{"role": "user", "content": "hi"}], llm_args()) == '{"ok": 1}'


def test_call_llm_403_is_actionable(monkeypatch):
    import requests
    monkeypatch.setattr(requests, "post", lambda *a, **k: _Resp(403, text="Forbidden"))
    with pytest.raises(SystemExit) as ei:
        e.call_llm([{"role": "user", "content": "hi"}], llm_args())
    assert "network policy" in str(ei.value)


def test_call_llm_bad_model_is_actionable(monkeypatch):
    import requests
    monkeypatch.setattr(requests, "post", lambda *a, **k: _Resp(400, text="model not found"))
    with pytest.raises(SystemExit) as ei:
        e.call_llm([{"role": "user", "content": "hi"}], llm_args(model="bogus/model"))
    assert "model" in str(ei.value).lower()


def test_call_llm_retries_then_succeeds(monkeypatch):
    import requests
    seq = [_Resp(503, text="busy"), _Resp(200, {"choices": [{"message": {"content": "ok"}}]})]
    monkeypatch.setattr(requests, "post", lambda *a, **k: seq.pop(0))
    monkeypatch.setattr(e.time, "sleep", lambda *_: None)
    assert e.call_llm([{"role": "user", "content": "hi"}], llm_args(retries=3)) == "ok"


def test_upload_filing_builds_request(monkeypatch):
    import requests
    captured = {}
    def fake_post(url, headers=None, json=None, timeout=None):
        captured.update(url=url, headers=headers, json=json)
        return _Resp(200, {"id": "f1"})
    monkeypatch.setattr(requests, "post", fake_post)
    out = e.upload_filing({"x": 1}, SimpleNamespace(api_url="https://qscreen.app/", token="tok"))
    assert out == {"id": "f1"}
    assert captured["url"] == "https://qscreen.app/api/v1/ingest/filing"
    assert captured["headers"]["Authorization"] == "Bearer tok"


# ── comparatives (prior-year column) ─────────────────────────────────────────

def test_validate_accepts_comparatives():
    f = good_filing()
    f["statements"][0]["line_items"][0]["comparatives"] = [{"period_label": "2022", "value": 9}]
    assert e.validate_filing(f) == []


def test_validate_rejects_bad_comparatives():
    f = good_filing()
    f["statements"][0]["line_items"][0]["comparatives"] = [{"value": 9}]   # no period_label
    assert any("comparatives" in p for p in e.validate_filing(f))
    f["statements"][0]["line_items"][0]["comparatives"] = "2022:9"          # not a list
    assert any("comparatives" in p for p in e.validate_filing(f))


def test_normalize_folds_prior_value_alias():
    d = {"metadata": {}, "audit": {}, "notes": [], "extraction_quality": {},
         "statements": [{"type": "income_statement", "verbatim_text": "x", "line_items": [
             {"account_code": "IS_REVENUE", "label_verbatim": "Revenue", "value": 10,
              "prior_value": 8, "prior_period_label": "2022"}]}]}
    n = e.normalize_filing(d)
    comps = n["statements"][0]["line_items"][0]["comparatives"]
    assert comps == [{"period_label": "2022", "value": 8}]


def test_merge_carries_comparatives_across_overlap():
    w1 = e.empty_filing(); w2 = e.empty_filing()
    w1["statements"].append({"type": "income_statement", "verbatim_text": "longer block ...........",
        "line_items": [{"label_verbatim": "Revenue", "value": 10, "account_code": "IS_REVENUE"}]})
    w2["statements"].append({"type": "income_statement", "verbatim_text": "x",
        "line_items": [{"label_verbatim": "Revenue", "value": 10, "account_code": "IS_REVENUE",
                        "comparatives": [{"period_label": "2022", "value": 8}]}]})
    st = e.merge_filings([w1, w2])["statements"][0]
    assert st["line_items"][0]["comparatives"] == [{"period_label": "2022", "value": 8}]


def test_flatten_includes_prior_columns():
    assert "prior_value" in e.EXPORT_COLUMNS and "prior_period_label" in e.EXPORT_COLUMNS
    f = good_filing()
    f["statements"][0]["line_items"][0]["comparatives"] = [{"period_label": "2022", "value": 9}]
    row = e.flatten_line_items(f)[0]
    assert row["prior_period_label"] == "2022" and row["prior_value"] == 9


# ── typed segments[] (contract) ──────────────────────────────────────────────

def test_empty_filing_has_segments_list():
    assert e.empty_filing()["segments"] == []


def test_validate_accepts_valid_segment():
    f = good_filing()
    f["segments"] = [{"dimension": "geography", "name": "Turkey", "currency": "TRY",
                      "metrics": {"revenue": 1}}]
    assert e.validate_filing(f) == []


def test_validate_rejects_bad_segment():
    f = good_filing()
    f["segments"] = [{"dimension": "planet", "name": "Mars"}]
    assert any("dimension" in p for p in e.validate_filing(f))
    f["segments"] = [{"dimension": "geography", "name": ""}]
    assert any("name" in p for p in e.validate_filing(f))


def test_normalize_coerces_segment_dimension_aliases():
    d = {"metadata": {}, "audit": {}, "notes": [], "extraction_quality": {}, "statements": [],
         "segments": [{"dimension": "Geographical", "name": "Egypt"},
                      {"dimension": "operating segment", "name": "Retail"},
                      {"dimension": "subsidiary", "name": "QNB Finansbank"},
                      {"name": "no dimension"}]}
    n = e.normalize_filing(d)
    dims = [s["dimension"] for s in n["segments"]]
    assert dims == ["geography", "business_line", "legal_entity", "business_line"]


def test_merge_unions_segments_across_windows():
    w1 = e.empty_filing(); w2 = e.empty_filing()
    w1["segments"] = [{"dimension": "geography", "name": "Qatar", "period_label": "2023",
                       "metrics": {"revenue": 100}}]
    w2["segments"] = [{"dimension": "geography", "name": "Turkey", "period_label": "2023",
                       "metrics": {"revenue": 50}, "verbatim_text": "Turkey segment ..."}]
    seg = e.merge_filings([w1, w2])["segments"]
    assert {s["name"] for s in seg} == {"Qatar", "Turkey"}


# ── profile-aware prompting (Qatar context injection) ────────────────────────

def test_qatar_context_empty_for_none():
    assert e._qatar_context(None) == ""


def test_system_prompt_injects_profile_context():
    import qatar
    sp = e._system_prompt("conventional_bank", False, qatar.profile_for_year("QNBK", 2016))
    assert "QATAR ANALYST CONTEXT" in sp
    assert "Turkey" in sp and "Basel III" in sp          # event timeline in force by 2016
    assert "comparatives" in sp                          # comparatives instruction present


def test_system_prompt_pre_acquisition_year_omits_segment():
    import qatar
    sp = e._system_prompt("conventional_bank", False, qatar.profile_for_year("QNBK", 2012))
    assert "QNB Finansbank" not in sp                    # Turkey not acquired until 2016


def test_system_prompt_without_profile_is_clean():
    sp = e._system_prompt("islamic_bank", False, None)
    # The segments rule mentions the context by name, but no per-company block is injected.
    assert "pre-loaded knowledge about THIS specific company" not in sp


# ── "both outputs": analysis artifacts + upload fold ─────────────────────────

def test_build_analysis_artifacts():
    from types import SimpleNamespace
    art = e.build_analysis_artifacts(good_filing(), SimpleNamespace(symbol="QNBK", _profile=None))
    assert "analysis" in art and "valuation" in art
    assert "ratios" in art["analysis"]


def test_upload_filing_folds_analysis_additively(monkeypatch):
    from types import SimpleNamespace

    class _Resp:
        def raise_for_status(self): pass
        def json(self): return {"ok": 1}
    cap = {}
    monkeypatch.setattr("requests.post",
                        lambda url, headers, json, timeout: cap.update(body=json) or _Resp())
    f = good_filing()
    args = SimpleNamespace(api_url="http://x", token="t")
    e.upload_filing(f, args, {"a": 1})
    assert cap["body"].get("analysis") == {"a": 1} and "metadata" in cap["body"]   # additive
    e.upload_filing(f, args)                       # default: plain filing, no analysis key
    assert "analysis" not in cap["body"]


# ── self-test entrypoint still green ─────────────────────────────────────────

def test_self_test_passes():
    assert e.run_self_test() == 0

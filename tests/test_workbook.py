"""Tests for the Excel financial-transcript workbook."""
from __future__ import annotations

import io

from openpyxl import load_workbook

import qscreen_workbook as wb


def _li(code, label, value, prior, depth=0, sub=False, note=None):
    return {"account_code": code, "label_verbatim": label, "value": value, "depth": depth,
            "is_subtotal": sub, "note_ref": note,
            "comparatives": [{"period_label": "2022", "value": prior}]}


def _filing():
    return {"metadata": {"symbol": "QNBK", "company_name": "Qatar National Bank",
                         "sector": "conventional_bank", "fiscal_year": 2023, "fiscal_period": "FY",
                         "currency": "QAR", "unit_scale": 1000, "reporting_framework": "IFRS",
                         "consolidated": True, "source_file": "qnb.pdf"},
            "audit": {"auditor_name": "KPMG", "opinion_type": "unqualified"},
            "statements": [
                {"type": "income_statement", "title": "Income Statement", "period_label": "2023",
                 "verbatim_text": "x", "line_items": [
                     _li("IS_NET_INTEREST", "Net interest income", 27800, 25600),
                     _li("IS_NET_INCOME", "Profit for the year", 15502, 14347, sub=True, note="10")]},
                {"type": "balance_sheet", "title": "Financial Position", "period_label": "2023",
                 "verbatim_text": "x", "line_items": [
                     _li("BS_TOTAL_ASSETS", "Total assets", 1200000, 1150000, sub=True),
                     _li("BS_LOANS", "Loans and advances", 830000, 810000, depth=1)]}],
            "segments": [{"dimension": "geography", "name": "Turkey", "currency": "TRY",
                          "period_label": "2023", "metrics": {"revenue": 8000, "net_profit": 600}}],
            "notes": [{"number": "1", "title": "Basis of preparation",
                       "category": "accounting_policies", "verbatim_text": "IFRS ..."}]}


def _load(filing, filings=None):
    return load_workbook(io.BytesIO(wb.workbook_bytes(filing, filings)))


def test_workbook_has_all_sheets():
    w = _load(_filing())
    assert {"Summary", "Income Statement", "Balance Sheet", "Financials (multi-year)",
            "Segments", "Notes"} <= set(w.sheetnames)
    assert w.sheetnames[0] == "Summary"


def test_statement_sheet_has_period_columns_and_numeric_cells():
    ist = _load(_filing())["Income Statement"]
    assert [c.value for c in ist[1]] == ["Label", "Code", "Note", "2023", "2022"]
    # IS_NET_INCOME row: current 15502, prior 14347, both numeric
    row = next(r for r in ist.iter_rows(min_row=2) if r[1].value == "IS_NET_INCOME")
    assert row[3].value == 15502 and row[4].value == 14347
    assert isinstance(row[3].value, (int, float)) and isinstance(row[4].value, (int, float))


def test_multiyear_grid_orders_codes_and_fills_years():
    my = _load(_filing())["Financials (multi-year)"]
    assert [c.value for c in my[1]] == ["Metric", "Code", "2022", "2023"]
    codes = [my.cell(r, 2).value for r in range(2, my.max_row + 1)]
    # grouped by statement (IS_ before BS_), alphabetical by canonical code within a group
    assert codes == ["IS_NET_INCOME", "IS_NET_INTEREST", "BS_LOANS", "BS_TOTAL_ASSETS"]
    ni = next(r for r in my.iter_rows(min_row=2) if r[1].value == "IS_NET_INCOME")
    assert ni[2].value == 14347 and ni[3].value == 15502


def test_segments_and_notes_sheets():
    w = _load(_filing())
    seg = w["Segments"]
    assert seg[1][0].value == "Dimension" and any(c.value == "Turkey" for c in seg[2])
    notes = w["Notes"]
    assert [c.value for c in notes[1]] == ["Number", "Title", "Category", "Text"]


def test_summary_sheet_unit_note():
    summ = _load(_filing())["Summary"]
    cells = [(summ.cell(r, 1).value, summ.cell(r, 2).value) for r in range(1, summ.max_row + 1)]
    assert ("Figures in", "thousands (×1,000)") in cells
    assert ("Symbol", "QNBK") in cells


def test_workbook_handles_thin_filing():
    thin = {"metadata": {"symbol": "X", "fiscal_year": 2023, "fiscal_period": "FY"},
            "statements": [{"type": "income_statement", "title": "IS", "period_label": "2023",
                            "verbatim_text": "x", "line_items": []}]}
    w = _load(thin)
    assert "Summary" in w.sheetnames and "Income Statement" in w.sheetnames    # no crash


def test_workbook_bytes_is_valid_xlsx():
    data = wb.workbook_bytes(_filing())
    assert data[:2] == b"PK" and len(data) > 2000


def _year_filing(year, ni):
    return {"metadata": {"symbol": "QNBK", "fiscal_year": year, "fiscal_period": "FY", "currency": "QAR"},
            "statements": [{"type": "income_statement", "title": "IS", "period_label": str(year),
                            "verbatim_text": "x", "line_items": [
                                {"account_code": "IS_NET_INCOME", "label_verbatim": "Profit", "value": ni,
                                 "comparatives": [{"period_label": str(year - 1), "value": ni - 1000}]}]}]}


def test_multi_filing_workbook_spans_all_years():
    f22, f23 = _year_filing(2022, 14000), _year_filing(2023, 15000)
    my = _load(f23, [f22, f23])["Financials (multi-year)"]
    headers = [c.value for c in my[1]]
    assert headers[:2] == ["Metric", "Code"] and {"2021", "2022", "2023"} <= set(headers)
    ni = next(r for r in my.iter_rows(min_row=2) if r[1].value == "IS_NET_INCOME")
    by_year = {my.cell(1, c).value: ni[c - 1].value for c in range(3, len(headers) + 1)}
    assert by_year["2023"] == 15000 and by_year["2022"] == 14000 and by_year["2021"] == 13000
